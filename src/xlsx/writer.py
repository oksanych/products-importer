from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook


PRODUCT_SHEET = "Export Products Sheet"


def write_xlsx_from_template(template_path: str, rows: list[dict], output_dir: str) -> str:
    if not rows:
        raise ValueError("No rows to write")

    workbook = load_workbook(template_path)
    sheet = workbook[PRODUCT_SHEET]
    headers = [cell.value for cell in sheet[1]]

    if sheet.max_row > 1:
        sheet.delete_rows(2, sheet.max_row - 1)

    for row_number, row_data in enumerate(rows, start=2):
        for column_number, header in enumerate(headers, start=1):
            if header in {"Назва_Характеристики", "Одиниця_виміру_Характеристики", "Значення_Характеристики"}:
                continue
            value = row_data.get(header, "")
            sheet.cell(row=row_number, column=column_number).value = value if value != "" else None
        _fill_characteristics(sheet, row_number, headers, row_data.get("_characteristics", []))

    output = Path(output_dir) / f"import-products-{rows[0]['_product_id']}.xlsx"
    output.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output)
    return str(output)


def _fill_characteristics(sheet, row_number: int, headers: list[str], characteristics: list[tuple[str, str, str]]) -> None:
    char_columns = [index + 1 for index, header in enumerate(headers) if header == "Назва_Характеристики"]
    for index, (name, unit, value) in enumerate(characteristics):
        if index >= len(char_columns):
            break
        column = char_columns[index]
        sheet.cell(row=row_number, column=column).value = name or None
        sheet.cell(row=row_number, column=column + 1).value = unit or None
        sheet.cell(row=row_number, column=column + 2).value = value or None

