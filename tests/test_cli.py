import subprocess
import sys
import tempfile
import unittest
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
SCRIPT = ROOT / "generate_import.py"
TEMPLATE = ROOT / "templates" / "export-template.xlsx"
FIXTURE = ROOT / "fixtures" / "modniy_3028043687.html"
URL = "https://modniy-shopping.com.ua/ua/p3028043687-product.html"


class CliTests(unittest.TestCase):
    def run_cli(self, *args):
        return subprocess.run(
            [sys.executable, str(SCRIPT), *args],
            cwd=ROOT,
            text=True,
            stdout=subprocess.PIPE,
            stderr=subprocess.PIPE,
        )

    def test_rejects_missing_color_id_with_clear_message(self):
        result = self.run_cli(URL, "--from-html", str(FIXTURE), "--template", str(TEMPLATE))

        self.assertEqual(result.returncode, 2)
        self.assertIn("Check color parameter", result.stderr)

    def test_rejects_color_flag_without_value_with_clear_message(self):
        result = self.run_cli(URL, "--from-html", str(FIXTURE), "--template", str(TEMPLATE), "--color-id")

        self.assertEqual(result.returncode, 2)
        self.assertIn("Check color parameter", result.stderr)

    def test_rejects_invalid_color_id_with_clear_message(self):
        for color_id in ["5", "ab"]:
            with self.subTest(color_id=color_id):
                result = self.run_cli(
                    URL,
                    "--from-html",
                    str(FIXTURE),
                    "--template",
                    str(TEMPLATE),
                    "--color-id",
                    color_id,
                )

                self.assertEqual(result.returncode, 2)
                self.assertIn("Check color parameter", result.stderr)

    def test_rejects_missing_import_price_with_clear_message(self):
        result = self.run_cli(URL, "--from-html", str(FIXTURE), "--template", str(TEMPLATE), "--color-id", "65")

        self.assertEqual(result.returncode, 2)
        self.assertIn("Check price parameter", result.stderr)

    def test_rejects_price_flag_without_value_with_clear_message(self):
        result = self.run_cli(
            URL,
            "--from-html",
            str(FIXTURE),
            "--template",
            str(TEMPLATE),
            "--color-id",
            "65",
            "--import-price",
        )

        self.assertEqual(result.returncode, 2)
        self.assertIn("Check price parameter", result.stderr)

    def test_rejects_invalid_import_price_with_clear_message(self):
        for import_price in ["0", "abc", "999.50"]:
            with self.subTest(import_price=import_price):
                result = self.run_cli(
                    URL,
                    "--from-html",
                    str(FIXTURE),
                    "--template",
                    str(TEMPLATE),
                    "--color-id",
                    "65",
                    "--import-price",
                    import_price,
                )

                self.assertEqual(result.returncode, 2)
                self.assertIn("Check price parameter", result.stderr)

    def test_rejects_missing_position_title_with_clear_message(self):
        result = self.run_cli(
            URL,
            "--from-html",
            str(FIXTURE),
            "--template",
            str(TEMPLATE),
            "--color-id",
            "65",
            "--import-price",
            "1000",
        )

        self.assertEqual(result.returncode, 2)
        self.assertIn("Check position title parameter", result.stderr)

    def test_rejects_missing_ukrainian_position_title_with_clear_message(self):
        result = self.run_cli(
            URL,
            "--from-html",
            str(FIXTURE),
            "--template",
            str(TEMPLATE),
            "--color-id",
            "65",
            "--import-price",
            "1000",
            "--position-title",
            "Купальник жіночий",
        )

        self.assertEqual(result.returncode, 2)
        self.assertIn("Check Ukrainian position title parameter", result.stderr)

    def test_rejects_position_title_that_exceeds_prom_limit_after_size_is_appended(self):
        result = self.run_cli(
            URL,
            "--from-html",
            str(FIXTURE),
            "--template",
            str(TEMPLATE),
            "--color-id",
            "65",
            "--import-price",
            "1000",
            "--position-title",
            "А" * 108,
            "--position-title-ukr",
            "Купальник жіночий",
        )

        self.assertEqual(result.returncode, 2)
        self.assertIn("Check position title parameter", result.stderr)

    def test_generates_xlsx_with_valid_color_id_import_price_and_position_titles(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            result = self.run_cli(
                URL,
                "--from-html",
                str(FIXTURE),
                "--template",
                str(TEMPLATE),
                "--output-dir",
                tmpdir,
                "--color-id",
                "65",
                "--import-price",
                "1000",
                "--position-title",
                "Купальник жіночий чорний",
                "--position-title-ukr",
                "Купальник жіночий чорний",
            )

            self.assertEqual(result.returncode, 0, result.stderr)
            output_path = Path(result.stdout.strip().removeprefix("Generated file: "))
            workbook = load_workbook(output_path, data_only=False)
            sheet = workbook["Export Products Sheet"]
            headers = [cell.value for cell in sheet[1]]
            product_code_col = headers.index("Код_товару") + 1
            price_col = headers.index("Ціна") + 1
            title_col = headers.index("Назва_позиції") + 1
            title_ukr_col = headers.index("Назва_позиції_укр") + 1

            self.assertEqual(sheet.cell(2, product_code_col).value, "MS5880065")
            self.assertEqual(sheet.cell(2, price_col).value, 1000)
            self.assertEqual(sheet.cell(2, price_col).data_type, "n")
            self.assertEqual(sheet.cell(2, title_col).value, "Купальник жіночий чорний 52")
            self.assertEqual(sheet.cell(2, title_ukr_col).value, "Купальник жіночий чорний 52")


if __name__ == "__main__":
    unittest.main()
