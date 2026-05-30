import sys
import tempfile
import unittest
from pathlib import Path

from openpyxl import load_workbook

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

from src.main import generate_import_file_from_html
from src.parser.modniy_shopping import parse_product_html
from src.xlsx.row_builder import build_xlsx_rows, numeric_crc32_group_id


ROOT = Path(__file__).resolve().parents[1]
FIXTURES = ROOT / "fixtures"
TEMPLATE = ROOT / "templates" / "export-template.xlsx"
EXPECTED_GROUP = {
    "Номер_групи": 31935910,
    "Назва_групи": "Купальники",
    "Назва_групи_укр": "Купальники",
    "Ідентифікатор_групи": "31935910",
    "Номер_батьківської_групи": None,
    "Ідентифікатор_батьківської_групи": None,
}


class XlsxContractTests(unittest.TestCase):
    def test_template_is_sanitized_for_public_repository(self):
        workbook = load_workbook(TEMPLATE, data_only=False)

        self.assertEqual(workbook.sheetnames, ["Export Products Sheet", "Export Groups Sheet"])
        product_sheet = workbook["Export Products Sheet"]
        product_headers = [cell.value for cell in product_sheet[1]]
        self.assertEqual(len(product_headers), 128)
        self.assertEqual(product_sheet.max_row, 1)

        group_sheet = workbook["Export Groups Sheet"]
        group_headers = [cell.value for cell in group_sheet[1]]
        header_to_index = {header: index + 1 for index, header in enumerate(group_headers) if header}
        self.assertEqual(group_sheet.max_row, 2)
        actual_group = {
            header: group_sheet.cell(2, header_to_index[header]).value
            for header in EXPECTED_GROUP
        }
        self.assertEqual(actual_group, EXPECTED_GROUP)

    def test_builds_rows_with_prom_variant_contract(self):
        html = (FIXTURES / "modniy_3064637917.html").read_text(encoding="utf-8")
        product = parse_product_html(html, "https://modniy-shopping.com.ua/ua/p3064637917-product.html")
        rows = build_xlsx_rows(
            product,
            color_id="65",
            position_title="Купальник жіночий чорний",
            position_title_ukr="Купальник жіночий чорний",
        )

        self.assertEqual(len(rows), 5)
        group_id = numeric_crc32_group_id("3064637917")
        self.assertTrue(0 < group_id <= 999_999_999)
        self.assertEqual({row["Код_товару"] for row in rows}, {"MS2512265"})
        self.assertEqual({row["ID_групи_різновидів"] for row in rows}, {str(group_id)})
        self.assertEqual({row["Унікальний_ідентифікатор"] for row in rows}, {""})
        self.assertEqual(len({row["Ідентифікатор_товару"] for row in rows}), 5)
        self.assertEqual(rows[0]["Ідентифікатор_товару"], "modniy-3064637917-48")
        self.assertEqual({row["Тип_товару"] for row in rows}, {"r"})
        self.assertEqual({row["Наявність"] for row in rows}, {"+"})
        self.assertEqual({row["Кількість"] for row in rows}, {"3"})
        self.assertEqual({row["Оптова_ціна"] for row in rows}, {""})
        self.assertEqual({row["Мінімальне_замовлення_опт"] for row in rows}, {""})
        self.assertEqual({row["Номер_групи"] for row in rows}, {31935910})
        self.assertEqual({row["Назва_групи"] for row in rows}, {"Купальники"})
        self.assertEqual({row["Ідентифікатор_групи"] for row in rows}, {"31935910"})
        self.assertEqual({row["Подарунки"] for row in rows}, {""})
        self.assertEqual({row["ID_Подарунків"] for row in rows}, {""})
        self.assertEqual({row["Супутні"] for row in rows}, {""})
        self.assertEqual({row["ID_Супутніх"] for row in rows}, {""})
        self.assertEqual([row["Назва_позиції"] for row in rows], [
            "Купальник жіночий чорний 48",
            "Купальник жіночий чорний 50",
            "Купальник жіночий чорний 52",
            "Купальник жіночий чорний 54",
            "Купальник жіночий чорний 56",
        ])
        self.assertEqual([row["Назва_позиції_укр"] for row in rows], [
            "Купальник жіночий чорний 48",
            "Купальник жіночий чорний 50",
            "Купальник жіночий чорний 52",
            "Купальник жіночий чорний 54",
            "Купальник жіночий чорний 56",
        ])
        self.assertEqual(rows[0]["Опис"].count("<img"), len(product.images))
        self.assertEqual(rows[0]["Опис_укр"].count("<img"), len(product.images))
        self.assertIn('<p style="text-align:center"><img', rows[0]["Опис"])
        self.assertIn('<p style="text-align:center"><img', rows[0]["Опис_укр"])
        self.assertLess(rows[0]["Опис"].index("<img"), rows[0]["Опис"].index("<table>"))
        self.assertNotIn("<img", rows[0]["Посилання_зображення"])

        names_by_row = [tuple(name for name, _unit, _value in row["_characteristics"]) for row in rows]
        self.assertEqual(len(set(names_by_row)), 1)
        size_values = [
            dict((name, value) for name, _unit, value in row["_characteristics"])["Розмір жіночого одягу (UA)"]
            for row in rows
        ]
        self.assertEqual(size_values, ["48", "50", "52", "54", "56"])

    def test_writes_xlsx_without_breaking_template_contract(self):
        html = (FIXTURES / "modniy_2999460479.html").read_text(encoding="utf-8")
        with tempfile.TemporaryDirectory() as tmpdir:
            output_path = generate_import_file_from_html(
                html,
                "https://modniy-shopping.com.ua/ua/p2999460479-product.html",
                template_path=str(TEMPLATE),
                output_dir=tmpdir,
                color_id="65",
                position_title="Купальник жіночий чорний",
                position_title_ukr="Купальник жіночий чорний",
            )

            workbook = load_workbook(output_path, data_only=False)
            self.assertEqual(workbook.sheetnames, ["Export Products Sheet", "Export Groups Sheet"])
            sheet = workbook["Export Products Sheet"]
            headers = [cell.value for cell in sheet[1]]
            self.assertEqual(len(headers), 128)
            self.assertEqual(sheet.max_row, 6)

            header_to_index = {header: index + 1 for index, header in enumerate(headers) if header}
            unique_id_col = header_to_index["Унікальний_ідентифікатор"]
            item_id_col = header_to_index["Ідентифікатор_товару"]
            variant_group_col = header_to_index["ID_групи_різновидів"]
            group_number_col = header_to_index["Номер_групи"]
            group_name_col = header_to_index["Назва_групи"]
            group_identifier_col = header_to_index["Ідентифікатор_групи"]
            wholesale_col = header_to_index["Оптова_ціна"]
            min_wholesale_col = header_to_index["Мінімальне_замовлення_опт"]

            variant_group_values = set()
            item_ids = set()
            for row_number in range(2, sheet.max_row + 1):
                self.assertIsNone(sheet.cell(row_number, unique_id_col).value)
                item_ids.add(sheet.cell(row_number, item_id_col).value)
                variant_group_values.add(sheet.cell(row_number, variant_group_col).value)
                self.assertEqual(sheet.cell(row_number, group_number_col).value, 31935910)
                self.assertEqual(sheet.cell(row_number, group_name_col).value, "Купальники")
                self.assertEqual(sheet.cell(row_number, group_identifier_col).value, "31935910")
                self.assertIsNone(sheet.cell(row_number, wholesale_col).value)
                self.assertIsNone(sheet.cell(row_number, min_wholesale_col).value)

            self.assertEqual(len(item_ids), 5)
            self.assertEqual(len(variant_group_values), 1)

            group_sheet = workbook["Export Groups Sheet"]
            group_headers = [cell.value for cell in group_sheet[1]]
            group_header_to_index = {header: index + 1 for index, header in enumerate(group_headers) if header}
            self.assertEqual(group_sheet.max_row, 2)
            generated_group = {
                header: group_sheet.cell(2, group_header_to_index[header]).value
                for header in EXPECTED_GROUP
            }
            self.assertEqual(generated_group, EXPECTED_GROUP)

    def test_writes_manual_price_override_as_numeric_value(self):
        html = (FIXTURES / "modniy_2999460479.html").read_text(encoding="utf-8")
        with tempfile.TemporaryDirectory() as tmpdir:
            output_path = generate_import_file_from_html(
                html,
                "https://modniy-shopping.com.ua/ua/p2999460479-product.html",
                template_path=str(TEMPLATE),
                output_dir=tmpdir,
                color_id="65",
                price_override=1000,
                position_title="Купальник жіночий чорний",
                position_title_ukr="Купальник жіночий чорний",
            )

            workbook = load_workbook(output_path, data_only=False)
            sheet = workbook["Export Products Sheet"]
            headers = [cell.value for cell in sheet[1]]
            price_col = headers.index("Ціна") + 1
            prices = [sheet.cell(row_number, price_col).value for row_number in range(2, sheet.max_row + 1)]
            data_types = [sheet.cell(row_number, price_col).data_type for row_number in range(2, sheet.max_row + 1)]

        self.assertEqual(prices, [1000, 1000, 1000, 1000, 1000])
        self.assertEqual(data_types, ["n", "n", "n", "n", "n"])

    def test_writes_manual_price_override_as_numeric_value(self):
        html = (FIXTURES / "modniy_2999460479.html").read_text(encoding="utf-8")
        with tempfile.TemporaryDirectory() as tmpdir:
            output_path = generate_import_file_from_html(
                html,
                "https://modniy-shopping.com.ua/ua/p2999460479-product.html",
                template_path=str(TEMPLATE),
                output_dir=tmpdir,
                color_id="65",
                price_override=1000,
                position_title="Купальник жіночий чорний",
                position_title_ukr="Купальник жіночий чорний",
            )

            workbook = load_workbook(output_path, data_only=False)
            sheet = workbook["Export Products Sheet"]
            headers = [cell.value for cell in sheet[1]]
            price_col = headers.index("Ціна") + 1
            prices = [sheet.cell(row_number, price_col).value for row_number in range(2, sheet.max_row + 1)]
            data_types = [sheet.cell(row_number, price_col).data_type for row_number in range(2, sheet.max_row + 1)]

        self.assertEqual(prices, [1000, 1000, 1000, 1000, 1000])
        self.assertEqual(data_types, ["n", "n", "n", "n", "n"])

    def test_writes_formatted_product_code_for_sku_with_suffix(self):
        html = (FIXTURES / "modniy_3028043687.html").read_text(encoding="utf-8")
        product = parse_product_html(html, "https://modniy-shopping.com.ua/ua/p3028043687-product.html")
        rows = build_xlsx_rows(
            product,
            color_id="65",
            position_title="Купальник жіночий чорний",
            position_title_ukr="Купальник жіночий чорний",
        )

        self.assertEqual(product.sku, "58800-25")
        self.assertEqual({row["Код_товару"] for row in rows}, {"MS5880065"})


if __name__ == "__main__":
    unittest.main()
